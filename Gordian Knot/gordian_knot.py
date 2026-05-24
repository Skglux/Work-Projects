"""
*Gordian Knot* -- Bonus Automation Script
=======================
Processes paired Report A / Report B Excel packages:
  - Unmerges cells, trims TLC (Three Letter Code)
  - Extracts Special TLC data from Report B and appends it to Report A
  - Deletes excluded TLC blocks
  - Removes duplicate day entries within each TLC block
  - Injects bonus calculation formulas
  - Saves outputs to a chosen folder
  - Optionally sends results by email

All sensitive values (TLC codes, bonus amounts, email credentials) are loaded
from config.json.
"""

import json
import os
import smtplib
import sys
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl.styles import Alignment


# ── Load config ───────────────────────────────────────────────────────────────
CONFIG_PATH = Path(__file__).parent / "config.json"

if not CONFIG_PATH.exists():
    raise SystemExit(
        "ERROR: config.json not found.\n"
    )

with open(CONFIG_PATH, "r", encoding="utf-8") as _f:
    _cfg = json.load(_f)

EMAIL_CONFIG    = _cfg["email"]
REPORT_A_CONFIG      = _cfg["report_a"]
REPORT_B_CONFIG   = _cfg["report_b"]


# ── UI bootstrap ──────────────────────────────────────────────────────────────
root = tk.Tk()
root.withdraw()

base_path = filedialog.askopenfilenames(title="Select Excel files to process")
if not base_path:
    raise SystemExit("No files selected – aborting.")

lst_path = list(base_path)
# sort(reverse=True) puts 'report_b' before 'report_a' alphabetically,
# which guarantees Special TLC data is extracted from Report B before Report A tries to use it.
lst_path.sort(reverse=True)

output_folder = filedialog.askdirectory(title="Select output folder for modified files")
if not output_folder:
    output_folder = os.path.dirname(lst_path[0])


# ── Helper: TLC block boundaries ─────────────────────────────────────────────
def _get_tlc_block_bounds(sheet, tlc_row: int) -> tuple[int, int]:
    """
    Returns (block_start_row, block_end_row) for the TLC block whose name
    sits at *tlc_row* in column 2.

    Assumed layout:
        tlc_row - 1  │ section / group header row
        tlc_row      │ TLC name  (col 2)
        tlc_row+1 …  │ sub-header rows
        tlc_row+4    │ last sub-header row
        tlc_row+5 …  │ data rows  (col 9 is the sentinel: empty col 9 → end of block)
    """
    block_start = tlc_row - 1
    data_start  = tlc_row + 5

    data_count = 0
    for r in range(data_start, 1001):
        if sheet.cell(row=r, column=9).value is None:
            break
        data_count += 1

    block_end = (data_start + data_count - 1) if data_count else (tlc_row + 4)
    return block_start, block_end


# ── Sheet processing functions ────────────────────────────────────────────────

def unmerge_cells(sheet):
    """Unmerges all merged regions that fall within columns 1-15."""
    for cell_group in list(sheet.merged_cells.ranges):
        if cell_group.min_col >= 1 and cell_group.max_col <= 15:
            sheet.unmerge_cells(str(cell_group))

    for row in sheet.iter_rows(min_col=1, max_col=15):
        for cell in row:
            cell.alignment = Alignment(wrap_text=False)

    return sheet


def trim_col2_values(sheet):
    """Strips leading/trailing whitespace from every TLC name in column 2."""
    for row_col2 in sheet.iter_rows(min_row=6, max_row=1000, min_col=2, max_col=2):
        cell = row_col2[0]
        if cell.value is not None and isinstance(cell.value, str):
            cell.value = cell.value.strip()
    return sheet


def bonus_index(sheet, BONUS):
    """Writes the bonus amount in a cell as a reference for each report"""
    sheet.cell(row=4, column=22).value = "BONUS:"
    sheet.cell(row=4, column=23).value = BONUS
    return sheet


def extract_special_tlc(data_dict: dict, sheet):
    """
    Reads Special TLC instruction-day rows (cols 9-13) from the Report B sheet and
    stores them in *data_dict* so Report A can append them later.
    """
    special_tlc = REPORT_B_CONFIG["special_tlc_code"]
    for row in sheet.iter_rows(min_row=6, max_row=1000, min_col=2, max_col=2):
        if row[0].value != special_tlc:
            continue

        data_start = row[0].row + 5
        for data_row in sheet.iter_rows(min_row=data_start, max_row=1000,
                                        min_col=9, max_col=13):
            if data_row[0].value is None:
                break
            for cell in data_row:
                data_dict[cell.column].append(cell.value)

        break  # stop after the first (and only) Special TLC block

    return data_dict, sheet


def append_special_tlc_insti_days(data_dict: dict, sheet):
    """
    Inserts rows extracted from the Report B Special TLC block into the matching
    block inside the Report A sheet.
    """
    if not any(data_dict.values()):
        print("  [WARN] data_dict is empty – nothing to append for Special TLC.")
        return data_dict, sheet

    data_points_count = len(next(iter(data_dict.values())))
    data_col          = list(data_dict.keys())
    special_tlc       = REPORT_A_CONFIG["special_tlc_code"]

    for row in sheet.iter_rows(min_row=6, max_row=1000, min_col=2, max_col=2):
        if row[0].value != special_tlc:
            continue

        data_start = row[0].row + 5
        sheet.insert_rows(idx=data_start, amount=data_points_count)

        for z in range(data_points_count):
            for col_idx in data_col:
                sheet.cell(row=data_start + z,
                           column=col_idx,
                           value=data_dict[col_idx][z])

        break  # stop after the first Special TLC block

    return data_dict, sheet


def tlc_data_input(sheet, tlc_modifiers: dict):
    """
    Populates the TLC modifier lookup table into hidden helper columns X/Y.
    Values come from config.json — no TLC codes are hardcoded in the script.
    """
    row_offset = 0
    for day_count_str, tlc_list in tlc_modifiers.items():
        day_count = int(day_count_str)
        for tlc_name in tlc_list:
            sheet.cell(row=1 + row_offset, column=24).value = tlc_name
            sheet.cell(row=1 + row_offset, column=25).value = day_count
            row_offset += 1
    return sheet


def headers(sheet):
    """Writes the seven calculation column headers starting at Q5."""
    new_headers = [
        "Inst. Day", "TLC", "Exception", "Count>",
        "Days count", "Paid count days", "Amount(EUR)",
    ]
    for i, name in enumerate(new_headers):
        sheet.cell(row=5, column=17 + i).value = name
    return sheet


def delete_excluded_tlc(sheet, EXLUDED_TLC: list):
    """
    Locates every TLC block whose name is in *EXLUDED_TLC*, collects the
    full row range for each block, then deletes them all bottom-to-top so
    row indices above each deletion stay valid.
    """
    blocks_to_delete: list[tuple[int, int]] = []

    row_idx = 6
    while row_idx < 1000:
        cell_tlc = sheet.cell(row=row_idx, column=2).value
        if isinstance(cell_tlc, str) and cell_tlc.strip() in EXLUDED_TLC:
            block_start, block_end = _get_tlc_block_bounds(sheet, row_idx)
            blocks_to_delete.append((block_start, block_end))
            row_idx = block_end + 2
        else:
            row_idx += 1

    for start, end in sorted(blocks_to_delete, reverse=True):
        row_count = end - start + 1
        print(f"  [DELETE] Rows {start}–{end} ({row_count} rows) – excluded TLC block")
        sheet.delete_rows(start, row_count)

    return sheet


def remove_duplicate_days(sheet):
    """
    Within each TLC data block, removes any row whose col-9 day value has
    already appeared in that same block (keeps the first occurrence).
    """
    current_tlc = None
    seen_days: set[str] = set()
    rows_to_delete: list[int] = []

    row_idx = 6
    while row_idx < 1000:
        col2_val = sheet.cell(row=row_idx, column=2).value
        col9_val = sheet.cell(row=row_idx, column=9).value

        if col2_val is not None and isinstance(col2_val, str) and len(col2_val.strip()) >= 2:
            current_tlc = col2_val.strip()
            seen_days   = set()
            row_idx    += 1
            continue

        if current_tlc is not None and col9_val is not None:
            day_key = str(col9_val).strip()
            if day_key in seen_days:
                rows_to_delete.append(row_idx)
            else:
                seen_days.add(day_key)

        row_idx += 1

    for r in sorted(rows_to_delete, reverse=True):
        print(f"  [DEDUP]  Removing duplicate day row {r}")
        sheet.delete_rows(r)

    return sheet


def fucntions_input(sheet, BONUS, DAYS_OVER_COUNT):
    """
    Injects Excel formulas into the calculation columns (Q–W) for rows 6-999.
    Column mapping:
        Q (17) = Inst. Day   R (18) = TLC         S (19) = Exception
        T (20) = Count>      U (21) = Days count   V (22) = Paid count days
        W (23) = Amount(EUR)
    """
    for r in range(6, 1000):
        sheet.cell(row=r, column=17).value = f"=LEFT(I{r},9)"
        sheet.cell(row=r, column=18).value = f"=TRIM(B{r})"
        sheet.cell(row=r, column=19).value = f"=COUNTIF($X$1:$X$9,R{r}) > 0"
        sheet.cell(row=r, column=20).value = (
            f"=IF(AND(S{r}=FALSE,LEN(R{r})>2),{DAYS_OVER_COUNT},"
            f"IF(S{r}=TRUE,INDEX($Y$1:$Y$9,MATCH(R{r},$X$1:$X$9,0)),\"\"))"
        )
        sheet.cell(row=r, column=21).value = (
            f"=IF(R{r}<>\"\",IF(Q{r+5}=\"\",0,"
            f"MATCH(TRUE,INDEX(Q{r+5}:$Q$1100=\"\",0),0)-1),\"\")"
        )
        sheet.cell(row=r, column=22).value = (
            f"=IF(AND(T{r}<>\"\",U{r}<>\"\"),"
            f"IF(U{r}-T{r}<=0,\"Falls Short\",U{r}-T{r}),\"\")"
        )
        sheet.cell(row=r, column=23).value = (
            f"=IF(AND(R{r}<>\"\",V{r}<>\"Falls Short\"),{BONUS}*V{r},\"\")"
        )
    return sheet


def hide_filter(sheet):
    """Hides helper columns X & Y and applies an auto-filter to Q5:W1000."""
    for col_letter in ("X", "Y"):
        sheet.column_dimensions[col_letter].hidden = True
    sheet.auto_filter.ref = "Q5:W1000"
    return sheet


# ── File-type detection ───────────────────────────────────────────────────────

def detect_file_type(path: str, sheet) -> str:
    """
    Determines whether a file is a 'Report A' or 'Report B' package.

    Priority:
      1. Content-based  – TLC names from config (excluded lists identify each type)
      2. Filename       – 'report_a' or 'report_b' in the filename
      3. UI dialog      – asks the user when both heuristics fail
    """
    report_a_markers = set(REPORT_A_CONFIG["excluded_tlc"])
    report_b_markers = set(REPORT_B_CONFIG["excluded_tlc"])

    for row in sheet.iter_rows(min_row=6, max_row=200, min_col=2, max_col=2):
        val = row[0].value
        if val and isinstance(val, str):
            val = val.strip()
            if val in report_b_markers:
                return "report_b"
            if val in report_a_markers:
                return "report_a"

    name = os.path.basename(path).lower()
    if "report_a" in name:
        return "report_a"
    if "report_b" in name:
        return "report_b"

    answer = messagebox.askquestion(
        "File Type",
        f"Cannot auto-detect the type of:\n{os.path.basename(path)}\n\n"
        "Is this a Report A file?\n(Click 'No' for Report B)",
    )
    return "report_a" if answer == "yes" else "report_b"


# ── Email helper ──────────────────────────────────────────────────────────────

def send_email(output_paths: list[str]) -> None:
    """
    Sends every file in *output_paths* as an attachment via SMTP.
    Only runs when config.json email.enabled is true.
    """
    if not EMAIL_CONFIG.get("enabled", False):
        return

    msg            = MIMEMultipart()
    msg["From"]    = EMAIL_CONFIG["sender"]
    msg["To"]      = ", ".join(EMAIL_CONFIG["recipients"])
    msg["Subject"] = EMAIL_CONFIG["subject"]
    msg.attach(MIMEText(EMAIL_CONFIG["body"], "plain"))

    for file_path in output_paths:
        with open(file_path, "rb") as fh:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(fh.read())
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f'attachment; filename="{os.path.basename(file_path)}"',
        )
        msg.attach(part)

    try:
        with smtplib.SMTP(EMAIL_CONFIG["smtp_server"], EMAIL_CONFIG["smtp_port"]) as srv:
            srv.starttls()
            srv.login(EMAIL_CONFIG["sender"], EMAIL_CONFIG["password"])
            srv.sendmail(EMAIL_CONFIG["sender"], EMAIL_CONFIG["recipients"], msg.as_string())
        print("[EMAIL] Files sent successfully.")
    except Exception as exc:
        print(f"[EMAIL ERROR] {exc}")


# ── Process configuration (built from config.json) ───────────────────────────

file_type_config = {
    "report_a": {
        "bonus_amount":   REPORT_A_CONFIG["bonus_amount"],
        "day_over_count": REPORT_A_CONFIG["day_over_count"],
        "excluded_tlc":   REPORT_A_CONFIG["excluded_tlc"],
        "tlc_modifiers":  REPORT_A_CONFIG["tlc_modifiers"],
        "process_flow": [
            unmerge_cells,
            trim_col2_values,
            bonus_index,
            append_special_tlc_insti_days,
            tlc_data_input,
            headers,
            delete_excluded_tlc,
            remove_duplicate_days,
            fucntions_input,
            hide_filter,
        ],
    },
    "report_b": {
        "bonus_amount":   REPORT_B_CONFIG["bonus_amount"],
        "day_over_count": REPORT_B_CONFIG["day_over_count"],
        "excluded_tlc":   REPORT_B_CONFIG["excluded_tlc"],
        "tlc_modifiers":  REPORT_B_CONFIG["tlc_modifiers"],
        "process_flow": [
            unmerge_cells,
            trim_col2_values,
            bonus_index,
            extract_special_tlc,
            tlc_data_input,
            headers,
            delete_excluded_tlc,
            remove_duplicate_days,
            fucntions_input,
            hide_filter,
        ],
    },
}

data_dict: dict[int, list] = {9: [], 10: [], 11: [], 12: [], 13: []}


# ── Main processing loop ──────────────────────────────────────────────────────

output_paths: list[str] = []

for path in lst_path:
    filename = os.path.basename(path)
    print(f"\n{'='*60}")
    print(f"[PROCESSING] {filename}")
    print(f"{'='*60}")

    wb    = openpyxl.load_workbook(path)
    sheet = wb.active

    file_type = detect_file_type(path, sheet)
    config    = file_type_config[file_type]

    BONUS           = config["bonus_amount"]
    DAYS_OVER_COUNT = config["day_over_count"]
    EXLUDED_TLC     = config["excluded_tlc"]
    TLC_MODIFIERS   = config["tlc_modifiers"]

    detected_label = file_type.upper().replace("REPORT_A", "Report A").replace("REPORT_B", "Report B")
    print(f"  Detected  : {detected_label}")
    print(f"  Bonus     : {BONUS} EUR  |  Threshold: >{DAYS_OVER_COUNT} days")

    for func in config["process_flow"]:
        fname = func.__name__
        print(f"  → {fname}")
        try:
            if fname == "bonus_index":
                sheet = func(sheet, BONUS)
            elif fname == "extract_special_tlc":
                data_dict, sheet = func(data_dict, sheet)
            elif fname == "append_special_tlc_insti_days":
                data_dict, sheet = func(data_dict, sheet)
            elif fname == "tlc_data_input":
                sheet = func(sheet, TLC_MODIFIERS)
            elif fname == "fucntions_input":
                sheet = func(sheet, BONUS, DAYS_OVER_COUNT)
            elif fname == "delete_excluded_tlc":
                sheet = func(sheet, EXLUDED_TLC)
            else:
                sheet = func(sheet)
        except Exception as exc:
            print(f"  [ERROR] {fname} raised: {exc}")
            raise

    out_name = f"Modified_{filename}"
    out_path = os.path.join(output_folder, out_name)
    wb.save(out_path)
    output_paths.append(out_path)
    print(f"  [SAVED]  → {out_path}")

send_email(output_paths)

print(f"\n{'='*60}")
print("[DONE] All files processed successfully.")
print(f"Output folder: {output_folder}")
print(f"{'='*60}")
