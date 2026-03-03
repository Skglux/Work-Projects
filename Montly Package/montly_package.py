import openpyxl
from openpyxl.styles import Alignment

xl = openpyxl.load_workbook("0126packageraw.xlsx")

sheet = xl.active

merged_ranges = list(sheet.merged_cells.ranges)

for cell_group in merged_ranges:
    if cell_group.min_col >= 1 and cell_group.max_col <= 15:
        sheet.unmerge_cells(str(cell_group))

for row in sheet.iter_rows(min_col=1,max_col=15):
    for cell in row:
        cell.alignment = Alignment(wrap_text=False) 


tlc_exceptions = ["ABC", "DEF", "GHI", "JKL", "MNO", "PQR", "STU", "VWX", "YZA"]

days_count_exceptions = [1, 2, 3, 4, 5, 6, 7, 8, 9]

sheet.cell(row=4, column=22).value = "Index XX,XX ="
sheet.cell(row=4, column=23).value = 150 

for itlc, val in enumerate(tlc_exceptions):
    sheet.cell(row=1+itlc, column=24).value = val
for idays, days_count_exceptions in enumerate(days_count_exceptions):
    sheet.cell(row=1+idays, column=25).value = days_count_exceptions

new_headers = ["Days", "TLC", "Exception", "Count>", "Days count", "Paid count days", "Amount(EUR)"]
start_col = 17
for i, header_name in enumerate(new_headers):
    sheet.cell(row=5, column=start_col + i).value = header_name


#Days Column
for row_left in range(6, 1000):
    formulaleft = f'=LEFT(I{row_left},9)'
    sheet.cell(row=row_left, column=17).value = formulaleft

#TLC Column 
for row_trim in range(6, 1000):
    formulatrim = f'=TRIM(B{row_trim})'
    sheet.cell(row=row_trim, column=18).value = formulatrim

#Exceptions Column
for row_countif in range(6, 1000):
    formcountif = f'=COUNTIF($X$1:$X$9,R{row_countif}) > 0'
    sheet.cell(row=row_countif, column=19).value = formcountif

#Count> Column
for row_countover in range(6, 1000):
    formcountover = f'=IF(AND(S{row_countover}=FALSE,LEN(R{row_countover})>2),4,IF(S{row_countover}=TRUE,INDEX($Y$1:$Y$9,MATCH(R{row_countover},$X$1:$X$9,0)),""))'
    sheet.cell(row=row_countover, column=20).value = formcountover

#Days Count Columnn
for row_dayscount in range(6, 1000):
    formdayscount = f'=IF(R{row_dayscount}<>"",IF(Q{row_dayscount+5}="",0,MATCH(TRUE,INDEX(Q{row_dayscount+5}:$Q$1100="",0),0)-1),"")'
    sheet.cell(row=row_dayscount, column=21).value = formdayscount

#Paid count days Column
for row_paidcount in range(6, 1000):
    formpaidcount = f'=IF(AND(T{row_paidcount}<>"",U{row_paidcount}<>""),IF(U{row_paidcount}-T{row_paidcount}<=0,"Falls Short",U{row_paidcount}-T{row_paidcount}),"")'
    sheet.cell(row=row_paidcount, column=22).value = formpaidcount

#Amount(EUR) Column
for row_amount in range(6, 1000):
    formamount = f'=IF(AND(R{row_amount}<>"",V{row_amount}<>"Falls Short"),150*V{row_amount},"")'
    sheet.cell(row=row_amount, column=23).value = formamount

columns_to_hide = ['X','Y']
for col_h in columns_to_hide:
    sheet.column_dimensions[col_h].hidden=True

sheet.auto_filter.ref = "Q5:W1000"    

xl.save("0126unrawtest.xlsx")
